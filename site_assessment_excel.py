# site_assessment_excel.py

import os
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage # Used for image resizing/processing
import logging
from datetime import datetime

# --- Logging Configuration ---
logger = logging.getLogger(__name__)

# --- CONFIGURATION ---
GREEN = '125435' # Dark Green for headers
LIGHT_GREEN_HEADER = 'D3E5D3' # Lighter Green for detail labels
LIGHT_GREY = 'F0F0F0' 
ROW_HEIGHT_IN_POINTS = 100 # Standard row height to fit the embedded image

# Define a thin but visible border style (color changed from 'AAAAAA' to '666666')
THIN_BORDER = Border(
    left=Side(style='thin', color='666666'),
    right=Side(style='thin', color='666666'),
    top=Side(style='thin', color='666666'),
    bottom=Side(style='thin', color='666666')
)

# --- HELPER FUNCTIONS (No changes needed) ---

def resize_and_process_image(base64_img, max_width_pixels=120, max_height_pixels=120):
    """
    Decodes base64 image data, resizes it using PIL, and returns an openpyxl.drawing.image.Image object.
    The resulting image is temporarily saved to a BytesIO stream.
    """
    if not base64_img:
        return None

    try:
        if ',' in base64_img:
            base64_data = base64_img.split(',')[1]
        else:
            base64_data = base64_img
            
        image_data = base64.b64decode(base64_data)
        img_stream = BytesIO(image_data)
        
        img_pil = PILImage.open(img_stream)
        
        # Calculate new size to fit within constraints while maintaining aspect ratio
        width, height = img_pil.size
        ratio = min(max_width_pixels / width, max_height_pixels / height)
        new_width = int(width * ratio)
        new_height = int(height * ratio)
        
        img_resized = img_pil.resize((new_width, new_height))
        
        output_stream = BytesIO()
        save_format = 'PNG' if img_pil.format in ('PNG', None) else img_pil.format
        img_resized.save(output_stream, format=save_format if save_format else 'PNG')
        output_stream.seek(0)
        
        return ExcelImage(output_stream)
        
    except Exception as e:
        logger.error(f"Image processing error for Excel: {e}")
        return None


# --- MAIN GENERATOR FUNCTION ---

def generate_assessment_excel(assessment_info):
    """Generates a professional Excel (XLSX) file from QHSE inspection data using openpyxl."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Defect Report"
    
    defect_items = assessment_info.get('defect_items', [])
    
    # --- STYLES ---
    title_font = Font(bold=True, size=16, color=GREEN)
    title_alignment = Alignment(horizontal='center', vertical='center')

    label_font = Font(bold=True, size=10, color='000000')
    label_fill = PatternFill(start_color=LIGHT_GREEN_HEADER, end_color=LIGHT_GREEN_HEADER, fill_type='solid')
    detail_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # --- 1. GENERAL INFORMATION HEADER BLOCK ---
    
    start_row = 1
    
    # Set the widths for the detail columns (A-H)
    ws.column_dimensions['A'].width = 18 # Label 1 (Wider for labels)
    ws.column_dimensions['B'].width = 25 # Value 1
    ws.column_dimensions['C'].width = 10 # Spacer
    ws.column_dimensions['D'].width = 18 # Label 2
    ws.column_dimensions['E'].width = 25 # Value 2 (Will be merged E:H)
    
    # Row 1: Main Title
    ws.merge_cells(f'A{start_row}:H{start_row}')
    title_cell = ws[f'A{start_row}']
    title_cell.value = f"QHSE DEFECT REPORT - {assessment_info.get('project_name', 'N/A').upper()}"
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    ws.row_dimensions[start_row].height = 25
    start_row += 1

    # Row 2: Project Name & Date
    r2 = start_row
    ws[f'A{r2}'].value = 'Project Name:'
    ws[f'B{r2}'].value = assessment_info.get('project_name', 'N/A')
    # FIX: Apply border to all cells in the merged range B:C
    ws[f'C{r2}'].border = THIN_BORDER 
    ws.merge_cells(f'B{r2}:C{r2}')
    
    ws[f'D{r2}'].value = 'Date of Visit:'
    ws[f'E{r2}'].value = assessment_info.get('date_of_visit', 'N/A')
    # FIX: Apply border to all cells in the merged range E:H
    for col_letter in ['F', 'G', 'H']: 
        ws[f'{col_letter}{r2}'].border = THIN_BORDER 
    ws.merge_cells(f'E{r2}:H{r2}')
    start_row += 1

    # Row 3: Building Location & Assessor
    r3 = start_row
    ws[f'A{r3}'].value = 'Building Location:'
    ws[f'B{r3}'].value = assessment_info.get('building_location', 'N/A')
    # FIX: Apply border to all cells in the merged range B:C
    ws[f'C{r3}'].border = THIN_BORDER 
    ws.merge_cells(f'B{r3}:C{r3}') 
    
    ws[f'D{r3}'].value = 'Assessor Name:'
    ws[f'E{r3}'].value = assessment_info.get('assessor_name', 'N/A')
    # FIX: Apply border to all cells in the merged range E:H
    for col_letter in ['F', 'G', 'H']: 
        ws[f'{col_letter}{r3}'].border = THIN_BORDER 
    ws.merge_cells(f'E{r3}:H{r3}')
    start_row += 1

    # Apply styles to the detail block
    for row in range(2, start_row):
        ws.row_dimensions[row].height = 18
        for col in ['A', 'D']:
            cell = ws[f'{col}{row}']
            cell.font = label_font
            cell.fill = label_fill
            cell.border = THIN_BORDER
            cell.alignment = detail_alignment
        
        # Apply border and alignment to the starting cell of the merged value blocks
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'E{row}'].border = THIN_BORDER
        ws[f'E{row}'].alignment = Alignment(horizontal='left', vertical='center')
            
    # Add a spacer row before the main table
    start_row += 1
    ws.row_dimensions[start_row].height = 5
    
    # --- 2. DEFECT ITEMS TABLE START ---

    # Adjust the column widths for the main table
    ws.column_dimensions['A'].width = 6    
    ws.column_dimensions['B'].width = 18   
    ws.column_dimensions['C'].width = 30   
    ws.column_dimensions['D'].width = 15   
    ws.column_dimensions['E'].width = 10   
    ws.column_dimensions['F'].width = 45   
    ws.column_dimensions['G'].width = 30   
    ws.column_dimensions['H'].width = 12   

    header_row_num = start_row + 1
    headers = [
        'Item No.', 'Location', 'Photo', 'Category', 
        'Priority', 'Description', 'Recommendation', 'Other Photos'
    ]
    
    ws.row_dimensions[header_row_num].height = 30 
    
    # This loop applies the dark green background, white font, centering, and borders
    for col_num, header_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f'{col_letter}{header_row_num}']
        cell.value = header_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = THIN_BORDER # Applies the bordered look
        
    # --- 3. WRITE DATA ROWS ---
    
    data_font = Font(size=10)
    data_alignment = Alignment(vertical='top', wrap_text=True)
    
    start_data_row = header_row_num + 1
    
    for item_index, item in enumerate(defect_items, 0): # item_index starts at 0
        row_num = start_data_row + item_index
        
        # Set a fixed height for all data rows to accommodate the image
        ws.row_dimensions[row_num].height = ROW_HEIGHT_IN_POINTS 
        
        all_photos = item.get('photos', [])
        primary_photo_b64 = all_photos[0] if all_photos else None
        remaining_photo_count = max(0, len(all_photos) - 1)

        # Data for columns A, B, D, E, F, G, H
        row_data = [
            item_index + 1, # Item No. (A)
            item.get('defect_location', 'N/A'), # Location (B)
            '', # Photo Placeholder (C)
            item.get('defect_category', 'N/A'), # Category (D)
            item.get('defect_priority', 'N/A'), # Priority (E)
            item.get('defect_description', 'N/A'), # Description (F)
            item.get('defect_recommendation', 'N/A'), # Recommendation (G)
            remaining_photo_count # Other Photos (H)
        ]
        
        for col_num, value in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f'{col_letter}{row_num}']
            cell.value = value
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = THIN_BORDER # Ensures all data cells have a visible border
            
            # Apply light gray background to every other row
            if item_index % 2 == 1:
                 cell.fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type='solid')

        # --- EMBED PRIMARY PHOTO (Column C) ---
        if primary_photo_b64:
            img = resize_and_process_image(primary_photo_b64)
            if img:
                ws.add_image(img, f'C{row_num}')
                
        # Center align the "Photo" and "Other Photos" column content for neatness
        ws[f'C{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'H{row_num}'].alignment = Alignment(horizontal='center', vertical='center')


    # --- 4. FINALIZE AND RETURN ---
    project_name = assessment_info.get('project_name', 'Unknown').replace(' ', '_')
    ts = datetime.now().strftime('%Y%m%d_%H%M%S') 
    filename = f"QHSE_Defect_Report_{project_name}_{ts}.xlsx"
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer, filename